"""
Auto Excel export module.
Exports sensor data to Excel files every N minutes.
"""
import os
import threading
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sensor_data import get_greenhouse_data, get_openfield_data, get_device_list, get_alerts
from config import EXCEL_EXPORT_DIR

# Ensure export directory exists
os.makedirs(EXCEL_EXPORT_DIR, exist_ok=True)

# History buffer for time-series data
_export_history = []
_lock = threading.Lock()

# Style constants
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2D8A6E", end_color="2D8A6E", fill_type="solid")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _record_snapshot():
    """Record a sensor snapshot to history buffer"""
    now = datetime.now()
    gh = get_greenhouse_data()
    of = get_openfield_data()
    entry = {"timestamp": now}
    entry.update({f"gh_{k}": v for k, v in gh.items()})
    entry.update({f"of_{k}": v for k, v in of.items()})
    with _lock:
        _export_history.append(entry)


def export_to_excel() -> str:
    """
    Export current + historical data to an Excel file.
    Returns the file path.
    """
    now = datetime.now()
    filename = f"慧植本草_传感器数据_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(EXCEL_EXPORT_DIR, filename)

    wb = Workbook()

    # === Sheet 1: 当前数据概览 ===
    ws1 = wb.active
    ws1.title = "数据概览"
    ws1.merge_cells("A1:F1")
    ws1["A1"] = f"慧植本草 — 传感器数据概览 ({now.strftime('%Y-%m-%d %H:%M:%S')})"
    ws1["A1"].font = TITLE_FONT
    ws1["A1"].alignment = Alignment(horizontal="center")

    # Greenhouse data
    row = 3
    ws1.cell(row=row, column=1, value="=== 有棚区传感器数据 ===").font = Font(bold=True, size=11)
    row += 1
    gh_headers = ["参数", "当前值", "单位", "适宜范围"]
    gh_ranges = {
        "光照": "12000-20000 lux",
        "pH值": "5.5-7.0",
        "土壤湿度": "30-60%",
        "CO₂浓度": "400-800 ppm",
        "空气湿度": "50-65%",
        "空气温度": "22-30°C",
        "钾(K)": "20-40%",
        "磷(P)": "150-250 mg/kg",
        "NDVI": "0.6-0.9",
        "蒸腾速率": "3-6",
        "电导率": "0.8-2.0",
    }
    gh_units = {
        "光照": "lux", "pH值": "", "土壤湿度": "%", "CO₂浓度": "ppm",
        "空气湿度": "%", "空气温度": "°C", "钾(K)": "%", "磷(P)": "mg/kg",
        "NDVI": "", "蒸腾速率": "", "电导率": "mS/cm",
    }

    for col, h in enumerate(gh_headers, 1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    gh = get_greenhouse_data()
    gh_keys = ["light", "ph", "soilMoisture", "co2", "airHumidity", "airTemp",
               "potassium", "phosphorus", "ndvi", "transpiration", "ec"]
    gh_names = ["光照", "pH值", "土壤湿度", "CO₂浓度", "空气湿度", "空气温度",
                "钾(K)", "磷(P)", "NDVI", "蒸腾速率", "电导率"]

    for i, (key, name) in enumerate(zip(gh_keys, gh_names)):
        r = row + 1 + i
        ws1.cell(row=r, column=1, value=name).border = THIN_BORDER
        ws1.cell(row=r, column=2, value=gh[key]).border = THIN_BORDER
        ws1.cell(row=r, column=3, value=gh_units[name]).border = THIN_BORDER
        ws1.cell(row=r, column=4, value=gh_ranges.get(name, "")).border = THIN_BORDER

    # Open field data
    row = row + len(gh_keys) + 2
    ws1.cell(row=row, column=1, value="=== 无棚区传感器数据 ===").font = Font(bold=True, size=11)
    row += 1
    for col, h in enumerate(["参数", "当前值", "单位"], 1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    of = get_openfield_data()
    of_data = [
        ("空气温度", of["airTemp"], "°C"),
        ("空气湿度", of["airHumidity"], "%"),
        ("CO₂浓度", of["co2"], "ppm"),
        ("土壤湿度", of["soilMoisture"], "%"),
        ("降雨量", of["rainfall"], "mm"),
        ("风速", of["windSpeed"], "m/s"),
        ("风向", of["windDir"], ""),
    ]
    for i, (name, val, unit) in enumerate(of_data):
        r = row + 1 + i
        ws1.cell(row=r, column=1, value=name).border = THIN_BORDER
        ws1.cell(row=r, column=2, value=val).border = THIN_BORDER
        ws1.cell(row=r, column=3, value=unit).border = THIN_BORDER

    # Auto-adjust column widths
    for col_idx, col in enumerate(ws1.columns, 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws1.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # === Sheet 2: 设备状态 ===
    ws2 = wb.create_sheet("设备状态")
    ws2.cell(row=1, column=1, value="设备运行状态").font = TITLE_FONT
    ws2.merge_cells("A1:E1")

    device_headers = ["设备名称", "状态", "功率/参数", "类别"]
    for col, h in enumerate(device_headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    devices = get_device_list()
    for i, d in enumerate(devices):
        r = 4 + i
        ws2.cell(row=r, column=1, value=d["name"]).border = THIN_BORDER
        status_text = {"on": "运行中", "off": "已关闭", "running": "运行中"}.get(d["status"], d["status"])
        ws2.cell(row=r, column=2, value=status_text).border = THIN_BORDER
        ws2.cell(row=r, column=3, value=d["value"]).border = THIN_BORDER
        ws2.cell(row=r, column=4, value=d["category"]).border = THIN_BORDER

    for col_idx, col in enumerate(ws2.columns, 1):
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws2.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    # === Sheet 3: 告警记录 ===
    ws3 = wb.create_sheet("告警记录")
    ws3.cell(row=1, column=1, value="告警记录").font = TITLE_FONT
    ws3.merge_cells("A1:E1")

    alert_headers = ["时间", "类型", "告警信息", "详细描述", "严重程度"]
    for col, h in enumerate(alert_headers, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    alerts = get_alerts()
    severity_colors = {
        "warning": "FFA500",
        "danger": "FF3366",
        "info": "4ECDC4",
    }
    for i, a in enumerate(alerts):
        r = 4 + i
        ws3.cell(row=r, column=1, value=a["time"]).border = THIN_BORDER
        ws3.cell(row=r, column=2, value=a["type"]).border = THIN_BORDER
        ws3.cell(row=r, column=3, value=a["message"]).border = THIN_BORDER
        ws3.cell(row=r, column=4, value=a["desc"]).border = THIN_BORDER
        cell = ws3.cell(row=r, column=5, value=a["severity"])
        cell.border = THIN_BORDER
        if a["severity"] in severity_colors:
            cell.font = Font(color=severity_colors[a["severity"]])

    for col_idx, col in enumerate(ws3.columns, 1):
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws3.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    # === Sheet 4: 历史趋势 ===
    with _lock:
        history = list(_export_history)

    if history:
        ws4 = wb.create_sheet("历史趋势")
        ws4.cell(row=1, column=1, value="传感器历史数据趋势").font = TITLE_FONT
        ws4.merge_cells("A1:D1")

        trend_headers = ["时间", "棚内温度", "棚内湿度", "CO₂浓度"]
        for col, h in enumerate(trend_headers, 1):
            cell = ws4.cell(row=3, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER

        for i, entry in enumerate(history[-200:]):  # Last 200 records
            r = 4 + i
            ws4.cell(row=r, column=1, value=entry["timestamp"].strftime("%H:%M:%S")).border = THIN_BORDER
            ws4.cell(row=r, column=2, value=entry.get("gh_airTemp", "")).border = THIN_BORDER
            ws4.cell(row=r, column=3, value=entry.get("gh_airHumidity", "")).border = THIN_BORDER
            ws4.cell(row=r, column=4, value=entry.get("gh_co2", "")).border = THIN_BORDER

        for col_idx, col in enumerate(ws4.columns, 1):
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws4.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    wb.save(filepath)
    return filepath


def auto_export_task():
    """Scheduled task: record snapshot + export Excel"""
    _record_snapshot()
    try:
        path = export_to_excel()
        print(f"[AutoExport] 导出完成: {path}")
    except Exception as e:
        print(f"[AutoExport] 导出失败: {e}")
